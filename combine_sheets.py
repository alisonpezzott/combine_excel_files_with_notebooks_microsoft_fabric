import pandas as pd
from openpyxl import load_workbook

excel_file = "Files/2025.xlsx"

def combine_sheets(excel_file):
    workbook = load_workbook(excel_file)
    sheet_names = workbook.sheetnames

    combined_df = pd.DataFrame()

    for sheet in sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet)

        df_melted = pd.melt(
                df,
                id_vars=[df.columns[0]],
                var_name="Date",
                value_name="Qty"
            )
        
        df_melted.dropna(subset=["Qty"], inplace=True)
        df_melted.rename(columns={df_melted.columns[0]: "ProductKey"}, inplace=True)
        df_melted["Date"] = pd.to_datetime(df_melted["Date"], dayfirst=True, errors="coerce").dt.date
        combined_df = pd.concat([combined_df, df_melted], ignore_index=True)
    
    print(combined_df)

if __name__ == '__main__':
    combine_sheets(excel_file)
